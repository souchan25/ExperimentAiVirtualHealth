from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, extract
from typing import List, Dict, Any
from uuid import UUID
import pandas as pd
import io
import os
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak, KeepTogether
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.colors import HexColor
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import logging

# CPSU Brand Colors
CPSU_GREEN = HexColor("#2E7D32")
CPSU_GREEN_LIGHT = HexColor("#4CAF50")
CPSU_GOLD = HexColor("#FFB300")
CPSU_TEXT = HexColor("#1A1A1A")

logger = logging.getLogger(__name__)

from ..database import get_db
from ..models import User, SymptomRecord, DepartmentStats, AuditLog, WellnessCheckin
from ..auth import get_current_user

router = APIRouter(prefix="/reports", tags=["reports"])

async def get_aggregated_report_data(db: AsyncSession):
    # 1. University-Wide Summary
    total_students_result = await db.execute(select(func.count(User.id)).where(User.role == "student"))
    total_students = total_students_result.scalar() or 0
    
    total_records_result = await db.execute(select(func.count(SymptomRecord.id)))
    total_consultations = total_records_result.scalar() or 0
    
    # Triage breakdown
    triage_result = await db.execute(select(SymptomRecord.severity, func.count(SymptomRecord.id)).group_by(SymptomRecord.severity))
    triage_rows = triage_result.all()
    triage_breakdown = {"Low": 0, "Moderate": 0, "High": 0, "Emergency": 0}
    for sev, count in triage_rows:
        if sev == 1: triage_breakdown["Low"] = count
        elif sev == 2: triage_breakdown["Moderate"] = count
        elif sev == 3: triage_breakdown["High"] = count
        else: triage_breakdown["Emergency"] = count
        
    # Status breakdown
    status_result = await db.execute(select(SymptomRecord.status, func.count(SymptomRecord.id)).group_by(SymptomRecord.status))
    status_breakdown = {row[0] or "pending": row[1] for row in status_result.all()}
    
    # --- Wellness Summary ---
    wellness_summary_res = await db.execute(
        select(
            func.avg(WellnessCheckin.stress_level),
            func.avg(WellnessCheckin.sleep_hours)
        )
    )
    avg_stress, avg_sleep = wellness_summary_res.first()
    
    mood_dist_res = await db.execute(
        select(WellnessCheckin.mood, func.count(WellnessCheckin.id))
        .group_by(WellnessCheckin.mood)
    )
    mood_distribution = {row[0]: row[1] for row in mood_dist_res.all() if row[0]}
    # ------------------------
    
    # 2. Per College/Department Breakdown
    from pydantic import BaseModel
    class DeptStatMock(BaseModel):
        department: str
        total_students: int
        students_with_symptoms: int
        percentage_with_symptoms: float
        top_diseases: list
        communicable_count: int
        non_communicable_count: int
        acute_count: int
        chronic_count: int
        referral_pending_count: int
        avg_wellness_stress: float
        dominant_mood: str

    # Bulk fetch department statistics
    dept_query = await db.execute(select(User.department).where(User.role == "student", User.department != "").distinct())
    departments = [row[0] for row in dept_query.all() if row[0]]
    
    # 1. Total students per department
    total_res = await db.execute(
        select(User.department, func.count(User.id))
        .where(User.role == "student", User.department != "")
        .group_by(User.department)
    )
    dept_totals = {row[0]: row[1] for row in total_res.all()}

    # 2. Students with symptoms per department
    symp_res = await db.execute(
        select(User.department, func.count(func.distinct(SymptomRecord.student_id)))
        .join(User, SymptomRecord.student_id == User.id)
        .where(User.department != "")
        .group_by(User.department)
    )
    dept_symptoms = {row[0]: row[1] for row in symp_res.all()}

    # 3. Pending referrals per department
    ref_res = await db.execute(
        select(User.department, func.count(SymptomRecord.id))
        .join(User, SymptomRecord.student_id == User.id)
        .where(User.department != "", SymptomRecord.status == "referred")
        .group_by(User.department)
    )
    dept_referrals = {row[0]: row[1] for row in ref_res.all()}

    # 4. Average wellness stress per department
    stress_res = await db.execute(
        select(User.department, func.avg(WellnessCheckin.stress_level))
        .join(User, WellnessCheckin.student_id == User.id)
        .where(User.department != "")
        .group_by(User.department)
    )
    dept_stress = {row[0]: row[1] for row in stress_res.all()}

    dept_stats = []
    for dept in departments:
        total_dept_students = dept_totals.get(dept, 0)
        dept_symptom_students = dept_symptoms.get(dept, 0)
        referral_pending = dept_referrals.get(dept, 0)
        avg_wellness_stress = round(dept_stress.get(dept) or 0.0, 1)

        percentage = (dept_symptom_students / total_dept_students * 100) if total_dept_students > 0 else 0
        
        dept_stats.append(DeptStatMock(
            department=dept,
            total_students=total_dept_students,
            students_with_symptoms=dept_symptom_students,
            percentage_with_symptoms=round(percentage, 2),
            top_diseases=[],
            communicable_count=0,
            non_communicable_count=0,
            acute_count=0,
            chronic_count=0,
            referral_pending_count=referral_pending,
            avg_wellness_stress=avg_wellness_stress,
            dominant_mood="None"
        ))
    
    # 3. Symptoms Frequency Analysis (Top 10)
    # This assumes symptoms are stored in a way we can extract. 
    # Since it's JSON, we might just aggregate predicted_disease as a proxy if symptoms are complex.
    symptoms_result = await db.execute(
        select(SymptomRecord.predicted_disease, func.count(SymptomRecord.id))
        .group_by(SymptomRecord.predicted_disease)
        .order_by(func.count(SymptomRecord.id).desc())
        .limit(10)
    )
    top_symptoms = [{"disease": row[0], "count": row[1]} for row in symptoms_result.all()]
    
    # Recurring symptoms flagging (students with 3+ consultations)
    recurring_result = await db.execute(
        select(SymptomRecord.student_id, func.count(SymptomRecord.id))
        .group_by(SymptomRecord.student_id)
        .having(func.count(SymptomRecord.id) >= 3)
    )
    recurring_student_ids = [row[0] for row in recurring_result.all()]
    
    # 4. Student Consultation Details
    logs_result = await db.execute(
        select(SymptomRecord, User.name, User.department, User.school_id)
        .join(User, SymptomRecord.student_id == User.id)
        .order_by(SymptomRecord.created_at.desc())
    )
    logs = []
    for row in logs_result.all():
        rec, name, dept, sid = row
        logs.append({
            "record_id": str(rec.id),
            "student_id": sid,
            "name": name,
            "department": dept,
            "date": rec.created_at.strftime("%Y-%m-%d %H:%M") if rec.created_at else "N/A",
            "triage": "Low" if rec.severity == 1 else "Moderate" if rec.severity == 2 else "High" if rec.severity == 3 else "Emergency",
            "status": rec.status,
            "disease": rec.predicted_disease
        })
        
    # 5. Staff Activity Summary
    # Cases reviewed (any record with status != 'pending' or having staff_notes)
    staff_activity = {}
    # This is a bit simplified, ideally we'd have a 'reviewed_by' field.
    # We'll use AuditLog to find 'update' actions on 'SymptomRecord'
    audit_result = await db.execute(
        select(User.name, func.count(AuditLog.id))
        .join(AuditLog, User.id == AuditLog.user_id)
        .where(AuditLog.model_name == "SymptomRecord", AuditLog.action == "update")
        .group_by(User.name)
    )
    for row in audit_result.all():
        staff_activity[row[0]] = {"reviews": row[1], "overrides": 0} # Overrides calculation needs more logic
        
    # Month-by-month triage trend
    trend_result = await db.execute(
        select(
            extract('month', SymptomRecord.created_at).label('month'),
            SymptomRecord.severity,
            func.count(SymptomRecord.id)
        )
        .group_by('month', SymptomRecord.severity)
        .order_by('month')
    )
    trend_data = {}
    for month, sev, count in trend_result.all():
        if month is None: continue
        m_name = datetime(2000, int(month), 1).strftime("%b")
        if m_name not in trend_data: trend_data[m_name] = {"Low": 0, "Moderate": 0, "High": 0, "Emergency": 0}
        label = "Low" if sev == 1 else "Moderate" if sev == 2 else "High" if sev == 3 else "Emergency"
        trend_data[m_name][label] = count

    # Staff Activity: Overrides calculation
    # Override = staff_diagnosis != predicted_disease and staff_diagnosis is not empty
    overrides_result = await db.execute(
        select(func.count(SymptomRecord.id))
        .where(SymptomRecord.staff_diagnosis != "", SymptomRecord.staff_diagnosis != SymptomRecord.predicted_disease)
    )
    total_overrides = overrides_result.scalar() or 0

    return {
        "summary": {
            "total_students": total_students,
            "total_consultations": total_consultations,
            "triage_breakdown": triage_breakdown,
            "status_breakdown": status_breakdown,
            "top_symptoms": top_symptoms,
            "recurring_count": len(recurring_student_ids),
            "total_overrides": total_overrides,
            "avg_stress": round(avg_stress or 0.0, 1),
            "avg_sleep": round(avg_sleep or 0.0, 1),
            "mood_distribution": mood_distribution
        },
        "department_stats": dept_stats,
        "logs": logs,
        "staff_activity": staff_activity,
        "trend_data": trend_data
    }

@router.get("/health-audit")
async def get_health_audit_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["staff", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
        
    data = await get_aggregated_report_data(db)
    
    # Simplify for frontend if needed, but original fields should remain
    return {
        "department_stats": [
            {
                "department": s.department,
                "total_students": s.total_students,
                "students_with_symptoms": s.students_with_symptoms,
                "percentage_with_symptoms": s.percentage_with_symptoms,
                "top_diseases": s.top_diseases,
                "communicable_count": s.communicable_count,
                "non_communicable_count": s.non_communicable_count,
                "acute_count": s.acute_count,
                "chronic_count": s.chronic_count,
                "referral_pending_count": s.referral_pending_count,
                "avg_wellness_stress": s.avg_wellness_stress
            } for s in data["department_stats"]
        ],
        "triage_breakdown": data["summary"]["triage_breakdown"],
        "total_records": data["summary"]["total_consultations"],
        "wellness_summary": {
            "avg_stress": data["summary"]["avg_stress"],
            "avg_sleep": data["summary"]["avg_sleep"],
            "mood_distribution": data["summary"]["mood_distribution"]
        }
    }

@router.get("/export/pdf")
async def export_pdf_report(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["staff", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
        
    data = await get_aggregated_report_data(db)
    
    os.makedirs("temp_reports", exist_ok=True)
    pdf_path = "temp_reports/health_audit_report.pdf"
    
    # Custom Page Template with Header/Footer
    class CPSUReportTemplate(SimpleDocTemplate):
        def __init__(self, filename, **kw):
            SimpleDocTemplate.__init__(self, filename, **kw)
            self.allowSplitting = 1

        def afterFlowable(self, flowable):
            "Registers page numbers"
            self.pageNum = self.canv.getPageNumber()

    def add_header_footer(canvas, doc):
        canvas.saveState()
        PageWidth, PageHeight = letter
        
        # Header area (within top margin)
        header_y = PageHeight - 0.4*inch
        
        # Logo
        logo_path = "d:/Expiremental/Assets/cpsu-logo.png"
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, doc.leftMargin, header_y - 0.2*inch, width=0.4*inch, height=0.4*inch, mask='auto')
        
        canvas.setFont('Helvetica-Bold', 10)
        canvas.setFillColor(CPSU_GREEN)
        canvas.drawString(doc.leftMargin + 0.5*inch, header_y, "CENTRAL PHILIPPINES STATE UNIVERSITY")
        
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.gray)
        canvas.drawString(doc.leftMargin + 0.5*inch, header_y - 0.15*inch, "University Health Services Clinic | Health Audit Report")
        
        # Header Line
        canvas.setStrokeColor(CPSU_GREEN)
        canvas.setLineWidth(0.5)
        canvas.line(doc.leftMargin, header_y - 0.3*inch, PageWidth - doc.rightMargin, header_y - 0.3*inch)
        
        # Footer
        footer_y = doc.bottomMargin
        canvas.line(doc.leftMargin, footer_y, PageWidth - doc.rightMargin, footer_y)
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.gray)
        canvas.drawString(doc.leftMargin, footer_y - 0.2*inch, f"Date Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas.drawRightString(PageWidth - doc.rightMargin, footer_y - 0.2*inch, f"Page {canvas.getPageNumber()}")
        
        canvas.restoreState()

    # Increased topMargin to 1.2 inch to accommodate header without overlap
    doc = CPSUReportTemplate(pdf_path, pagesize=letter, topMargin=1.2*inch, bottomMargin=0.8*inch)
    styles = getSampleStyleSheet()
    
    # Custom Styles
    styles.add(ParagraphStyle(
        name='CPSU_Title',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=CPSU_GREEN,
        alignment=TA_CENTER,
        spaceAfter=12
    ))
    
    styles.add(ParagraphStyle(
        name='CPSU_SubTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=CPSU_GOLD,
        alignment=TA_CENTER,
        spaceAfter=24
    ))

    styles.add(ParagraphStyle(
        name='CPSU_Heading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=CPSU_GREEN,
        spaceBefore=12,
        spaceAfter=6,
        borderPadding=(0, 0, 2, 0),
        borderWidth=0,
        borderColor=CPSU_GREEN
    ))

    styles.add(ParagraphStyle(
        name='Metric_Label',
        fontSize=9,
        textColor=colors.gray,
        alignment=TA_CENTER
    ))

    styles.add(ParagraphStyle(
        name='Metric_Value',
        fontSize=16,
        fontName='Helvetica-Bold',
        textColor=CPSU_GREEN,
        alignment=TA_CENTER
    ))

    elements = []
    
    # --- Title Page (Premium Design) ---
    elements.append(Spacer(1, 1.5*inch))
    logo_path = "d:/Expiremental/Assets/cpsu-logo.png"
    if os.path.exists(logo_path):
        img = Image(logo_path, 2*inch, 2*inch)
        img.hAlign = 'CENTER'
        elements.append(img)
    
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("CENTRAL PHILIPPINES STATE UNIVERSITY", styles['CPSU_Title']))
    elements.append(Paragraph("University Health Services Clinic", styles['CPSU_SubTitle']))
    elements.append(Spacer(1, 1.5*inch))
    
    elements.append(Paragraph("HEALTH AUDIT REPORT", styles['CPSU_Title']))
    elements.append(Paragraph(f"Period: Academic Year 2024-2025", styles['Normal'])) # Mocking period
    elements.append(Paragraph(f"Date Generated: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    elements.append(Paragraph(f"Generated By: {current_user.name}", styles['Normal']))
    
    elements.append(PageBreak())
    
    # Section 1: Executive Summary & Key Metrics
    elements.append(Paragraph("1. Executive Summary", styles['CPSU_Heading']))
    elements.append(Paragraph("This report provides an overview of health-related activities and trends within the university campus.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))

    # Metric Cards (Table-based)
    metric_data = [
        [
            Paragraph("TOTAL STUDENTS", styles['Metric_Label']),
            Paragraph("CONSULTATIONS", styles['Metric_Label']),
            Paragraph("AVG STRESS", styles['Metric_Label']),
            Paragraph("AVG SLEEP", styles['Metric_Label'])
        ],
        [
            Paragraph(str(data["summary"]["total_students"]), styles['Metric_Value']),
            Paragraph(str(data["summary"]["total_consultations"]), styles['Metric_Value']),
            Paragraph(f"{data['summary']['avg_stress']}/10", styles['Metric_Value']),
            Paragraph(f"{data['summary']['avg_sleep']} hrs", styles['Metric_Value'])
        ]
    ]
    
    metric_table = Table(metric_data, colWidths=[1.5*inch] * 4, hAlign='CENTER')
    metric_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(metric_table)
    elements.append(Spacer(1, 0.4*inch))

    # Triage Analysis (Chart)
    elements.append(Paragraph("Triage Severity Distribution", styles['Heading3']))
    
    # Body width is roughly 6.5 inches (468 pts). Center 400 pts drawing.
    d = Drawing(468, 200) 
    pc = Pie()
    pc.x = 184 # (468 - 100) / 2
    pc.y = 50
    pc.width = 100
    pc.height = 100
    
    triage_v = data["summary"]["triage_breakdown"]
    pc.data = [triage_v["Low"], triage_v["Moderate"], triage_v["High"], triage_v["Emergency"]]
    pc.labels = ['Low', 'Moderate', 'High', 'Emergency']
    pc.slices[0].fillColor = CPSU_GREEN
    pc.slices[1].fillColor = colors.orange
    pc.slices[2].fillColor = colors.red
    pc.slices[3].fillColor = colors.darkred
    
    d.add(pc)
    elements.append(d)
    elements.append(Spacer(1, 0.3*inch))

    # Existing Summary Table (Restyled)
    elements.append(Paragraph("University-Wide Summary Details", styles['Heading3']))
    summary_data = [
        ["Metric", "Value"],
        ["Low Priority Cases", triage_v["Low"]],
        ["Moderate Cases", triage_v["Moderate"]],
        ["High Priority Cases", triage_v["High"]],
        ["Emergency Calls", triage_v["Emergency"]],
        ["Recurring/Flagged Students", data["summary"]["recurring_count"]],
        ["AI Diagnosis Overrides", data["summary"]["total_overrides"]]
    ]
    t = Table(summary_data, colWidths=[3.5*inch, 1.5*inch], hAlign='CENTER')
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), CPSU_GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(t)
    elements.append(PageBreak())
    
    # Section 2: Department Breakdown
    elements.append(Paragraph("2. College/Department Breakdown", styles['CPSU_Heading']))
    
    # Bar Chart: Cases per Department
    dept_names = [d.department[:15] for d in data["department_stats"] if d.students_with_symptoms > 0]
    dept_counts = [d.students_with_symptoms for d in data["department_stats"] if d.students_with_symptoms > 0]
    
    if dept_counts:
        d_dept = Drawing(468, 200)
        bc = VerticalBarChart()
        bc.x = 84 # Center a 300 wide chart in 468 pts
        bc.y = 50
        bc.height = 125
        bc.width = 300
        bc.data = [dept_counts]
        bc.categoryAxis.categoryNames = dept_names
        bc.categoryAxis.labels.angle = 45
        bc.categoryAxis.labels.boxAnchor = 'ne'
        bc.bars[0].fillColor = CPSU_GREEN
        d_dept.add(bc)
        elements.append(d_dept)
        elements.append(Paragraph("Total Symptom Cases by Department", styles['Italic']))
        elements.append(Spacer(1, 0.3*inch))

    dept_data = [["College", "Symptom Cases", "Avg Stress", "Referrals"]]
    for d in data["department_stats"]:
        dept_data.append([d.department, d.students_with_symptoms, f"{d.avg_wellness_stress}/10", d.referral_pending_count])
    
    t_dept = Table(dept_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1*inch], hAlign='CENTER')
    t_dept.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), CPSU_GREEN),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 10),
    ]))
    elements.append(t_dept)
    
    # Section: Mood Distribution
    elements.append(Paragraph("3. Student Mood Distribution", styles['CPSU_Heading']))
    mood_items = list(data["summary"]["mood_distribution"].items())
    
    if mood_items:
        # Mini Bar Chart for mood
        d_mood = Drawing(468, 150)
        bc_mood = VerticalBarChart()
        bc_mood.x = 84
        bc_mood.y = 30
        bc_mood.height = 100
        bc_mood.width = 300
        bc_mood.data = [[count for _, count in mood_items]]
        bc_mood.categoryAxis.categoryNames = [mood for mood, _ in mood_items]
        bc_mood.bars[0].fillColor = CPSU_GOLD
        d_mood.add(bc_mood)
        elements.append(d_mood)
        
        mood_data = [["Mood", "Count"]]
        for mood, count in mood_items:
            mood_data.append([mood, count])
        
        t_mood = Table(mood_data, colWidths=[2*inch, 1*inch], hAlign='CENTER')
        t_mood.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('LINEBELOW', (0,0), (-1,0), 2, CPSU_GREEN),
        ]))
        elements.append(t_mood)
    else:
        elements.append(Paragraph("No mood data recorded in this period.", styles['Normal']))
    
    elements.append(PageBreak())
    
    # Section 4: Trends & Analysis
    elements.append(Paragraph("4. Triage Level Analysis & Trends", styles['CPSU_Heading']))
    trend_rows = [["Month", "Low", "Moderate", "High", "Emergency"]]
    for month, counts in data["trend_data"].items():
        trend_rows.append([month, counts["Low"], counts["Moderate"], counts["High"], counts["Emergency"]])
    
    if len(trend_rows) > 1:
        t_trend = Table(trend_rows, colWidths=[1.1*inch, 1*inch, 1*inch, 1*inch, 1.2*inch], hAlign='CENTER')
        t_trend.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('LINEBELOW', (0,0), (-1,0), 2, CPSU_GOLD),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]))
        elements.append(t_trend)
    else:
        elements.append(Paragraph("No trend data available for the selected period.", styles['Normal']))
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Section 4: Symptoms Frequency Analysis
    elements.append(Paragraph("5. Symptoms Frequency Analysis", styles['CPSU_Heading']))
    symptom_data = [["Predicted Condition", "Frequency / Count"]]
    for s in data["summary"]["top_symptoms"]:
        symptom_data.append([s["disease"], s["count"]])
    
    t_symp = Table(symptom_data, colWidths=[3.5*inch, 1.5*inch], hAlign='CENTER')
    t_symp.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), CPSU_GREEN),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
    ]))
    elements.append(t_symp)
    elements.append(Spacer(1, 0.3*inch))
    
    # Section 5: Staff Activity & Student Alerts
    elements.append(Paragraph("6. Staff Activity & System Alerts", styles['CPSU_Heading']))
    
    staff_rows = [["Staff Member", "Cases Reviewed"]]
    for name, act in data["staff_activity"].items():
        staff_rows.append([name, act["reviews"]])
    
    if len(staff_rows) > 1:
        t_staff = Table(staff_rows, colWidths=[3.5*inch, 1.5*inch], hAlign='CENTER')
        t_staff.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('LINEBELOW', (0,0), (-1,0), 2, colors.grey),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ALIGN', (1,0), (1,-1), 'CENTER'),
        ]))
        elements.append(t_staff)
    else:
        elements.append(Paragraph("No staff activity recorded.", styles['Normal']))

    # Signature Section
    elements.append(Spacer(1, 0.5*inch)) # Reduced spacer
    
    sig_data = [
        [Paragraph("<b>Certified Correct:</b>", styles['Normal']), Paragraph("<b>Approved By:</b>", styles['Normal'])],
        [Spacer(1, 0.4*inch), Spacer(1, 0.4*inch)], # Adjusted internal padding
        [Paragraph("__________________________", styles['Normal']), Paragraph("__________________________", styles['Normal'])],
        [Paragraph(f"<b>{current_user.name}</b>", styles['Normal']), Paragraph("<b>Clinic Head / Medical Officer</b>", styles['Normal'])],
        [Paragraph("University Clinic Staff", styles['Normal']), Paragraph("University Health Services", styles['Normal'])]
    ]
    
    sig_table = Table(sig_data, colWidths=[3*inch, 3*inch])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    # Wrap in KeepTogether to prevent splitting across pages
    elements.append(KeepTogether(sig_table))
    
    doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
    
    return FileResponse(pdf_path, filename="health_audit_report.pdf")

@router.get("/export/xlsx")
async def export_excel_report(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["staff", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
        
    try:
        data = await get_aggregated_report_data(db)
        
        output = io.BytesIO()
        
        # Define CPSU Colors
        CPSU_GREEN = "2E7D32"
        CPSU_GOLD = "FFB300"
        WHITE = "FFFFFF"
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # --- Sheet 1: Dashboard ---
            summary_df = pd.DataFrame([
                {"Metric": "Total Students", "Value": data["summary"]["total_students"]},
                {"Metric": "Total Consultations", "Value": data["summary"]["total_consultations"]},
                {"Metric": "Low Triage", "Value": data["summary"]["triage_breakdown"]["Low"]},
                {"Metric": "Moderate Triage", "Value": data["summary"]["triage_breakdown"]["Moderate"]},
                {"Metric": "High Triage", "Value": data["summary"]["triage_breakdown"]["High"]},
                {"Metric": "Emergency Triage", "Value": data["summary"]["triage_breakdown"]["Emergency"]},
                {"Metric": "Recurring Students", "Value": data["summary"]["recurring_count"]},
                {"Metric": "Avg Stress Level", "Value": data["summary"]["avg_stress"]},
                {"Metric": "Avg Sleep Hours", "Value": data["summary"]["avg_sleep"]}
            ])
            summary_df.to_excel(writer, index=False, sheet_name='Dashboard', startrow=4)
            
            workbook = writer.book
            ws_dash = writer.sheets['Dashboard']
            
            # Dashboard Header
            ws_dash.merge_cells('A1:C1')
            ws_dash['A1'] = "CENTRAL PHILIPPINES STATE UNIVERSITY"
            ws_dash['A1'].font = Font(size=14, bold=True, color=WHITE)
            ws_dash['A1'].fill = PatternFill(start_color=CPSU_GREEN, end_color=CPSU_GREEN, fill_type="solid")
            ws_dash['A1'].alignment = Alignment(horizontal="center")

            ws_dash.merge_cells('A2:C2')
            ws_dash['A2'] = "HEALTH AUDIT DASHBOARD SUMMARY"
            ws_dash['A2'].font = Font(size=12, bold=True, color=CPSU_GREEN)
            ws_dash['A2'].alignment = Alignment(horizontal="center")

            ws_dash['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws_dash['A3'].font = Font(italic=True)

            # Style Summary Table Headers
            for cell in ws_dash[5]:
                cell.font = Font(bold=True, color=WHITE)
                cell.fill = PatternFill(start_color=CPSU_GREEN, end_color=CPSU_GREEN, fill_type="solid")

            # PIE CHART: Triage Breakdown
            pie = PieChart()
            # Values are in B8:B11 (Low to Emergency) - Row index 8 to 11
            labels = Reference(ws_dash, min_col=1, min_row=8, max_row=11) 
            chart_data = Reference(ws_dash, min_col=2, min_row=8, max_row=11)
            pie.add_data(chart_data, titles_from_data=False)
            pie.set_categories(labels)
            pie.title = "Triage Severity Distribution"
            ws_dash.add_chart(pie, "E5")

            # BAR CHART: Mood Distribution (Place on Dashboard)
            mood_items = list(data["summary"]["mood_distribution"].items())
            if mood_items:
                start_row_mood = 16
                ws_dash.cell(row=start_row_mood, column=1, value="Mood")
                ws_dash.cell(row=start_row_mood, column=2, value="Count")
                for i, (mood, count) in enumerate(mood_items):
                    ws_dash.cell(row=start_row_mood+1+i, column=1, value=mood)
                    ws_dash.cell(row=start_row_mood+1+i, column=2, value=count)
                
                bar_mood = BarChart()
                mood_labels = Reference(ws_dash, min_col=1, min_row=start_row_mood+1, max_row=start_row_mood+len(mood_items))
                mood_data_ref = Reference(ws_dash, min_col=2, min_row=start_row_mood+1, max_row=start_row_mood+len(mood_items))
                bar_mood.add_data(mood_data_ref, titles_from_data=False)
                bar_mood.set_categories(mood_labels)
                bar_mood.title = "Student Mood Distribution"
                ws_dash.add_chart(bar_mood, "E20")

            # --- Sheet 2: Departmental ---
            dept_df = pd.DataFrame([
                {
                    "Department": d.department,
                    "Total Students": d.total_students,
                    "Cases": d.students_with_symptoms,
                    "Prevalence %": d.percentage_with_symptoms,
                    "Avg Stress": d.avg_wellness_stress,
                    "Referrals": d.referral_pending_count
                } for d in data["department_stats"]
            ])
            dept_df.to_excel(writer, index=False, sheet_name='DepartmentalAnalysis')
            ws_dept = writer.sheets['DepartmentalAnalysis']
            for cell in ws_dept[1]:
                cell.font = Font(bold=True, color=WHITE)
                cell.fill = PatternFill(start_color=CPSU_GREEN, end_color=CPSU_GREEN, fill_type="solid")
            
            # Bar Chart: Cases per Department
            if not dept_df.empty:
                bar_dept = BarChart()
                dept_labels = Reference(ws_dept, min_col=1, min_row=2, max_row=len(dept_df)+1)
                dept_data_ref = Reference(ws_dept, min_col=3, min_row=2, max_row=len(dept_df)+1)
                bar_dept.add_data(dept_data_ref, titles_from_data=False)
                bar_dept.set_categories(dept_labels)
                bar_dept.title = "Cases by College/Department"
                ws_dept.add_chart(bar_dept, "H2")

            # --- Sheet 3: Trends ---
            trend_list = []
            for month, counts in data["trend_data"].items():
                trend_list.append({"Month": month, **counts})
            
            trend_df = pd.DataFrame(trend_list)
            if not trend_df.empty:
                trend_df.to_excel(writer, index=False, sheet_name='Trends')
                ws_trend = writer.sheets['Trends']
                for cell in ws_trend[1]:
                    cell.font = Font(bold=True, color=WHITE)
                    cell.fill = PatternFill(start_color=CPSU_GOLD, end_color=CPSU_GOLD, fill_type="solid")
                
                # Line Chart: Trends
                line = LineChart()
                line_labels = Reference(ws_trend, min_col=1, min_row=2, max_row=len(trend_df)+1)
                line_data = Reference(ws_trend, min_col=2, max_col=5, min_row=1, max_row=len(trend_df)+1)
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(line_labels)
                line.title = "Monthly Triage Trends"
                ws_trend.add_chart(line, "G2")

            # --- Sheet 4: Raw Logs ---
            logs_df = pd.DataFrame(data["logs"])
            logs_df.to_excel(writer, index=False, sheet_name='ConsultationLogs')
            ws_logs = writer.sheets['ConsultationLogs']
            for cell in ws_logs[1]:
                cell.font = Font(bold=True, color=WHITE)
                cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

            # Auto-adjust column widths
            for sheetname in writer.sheets:
                ws = writer.sheets[sheetname]
                for col in ws.columns:
                    max_length = 0
                    column_idx = col[0].column # Integer index
                    col_letter = get_column_letter(column_idx)
                    for cell in col:
                        try:
                            if cell.value:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                        except: pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[col_letter].width = adjusted_width

        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=health_audit_report.xlsx"}
        )
    except Exception as e:
        logger.error(f"Excel Export Error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel report: {str(e)}")

@router.get("/referral/{record_id}/pdf")
async def export_referral_pdf(
    record_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["staff", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
        
    result = await db.execute(
        select(SymptomRecord, User.name, User.department, User.school_id)
        .join(User, SymptomRecord.student_id == User.id)
        .where(SymptomRecord.id == record_id)
    )
    row = result.first()
    
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
        
    record, student_name, student_dept, student_sid = row
    
    os.makedirs("temp_reports", exist_ok=True)
    pdf_path = f"temp_reports/referral_{record_id}.pdf"
    
    c = canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter
    
    # --- Letterhead ---
    logo_path = "d:/Expiremental/Assets/cpsu-logo.png"
    if os.path.exists(logo_path):
        c.drawImage(logo_path, 1*inch, height - 1.2*inch, width=0.8*inch, height=0.8*inch, mask='auto')
    
    c.setFillColor(CPSU_GREEN)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2*inch, height - 0.7*inch, "CENTRAL PHILIPPINES STATE UNIVERSITY")
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*inch, height - 0.9*inch, "University Health Services Clinic")
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.gray)
    c.drawString(2*inch, height - 1.05*inch, "Kabankalan City, Negros Occidental, Philippines | healthclinic@cpsu.edu.ph")
    
    c.setStrokeColor(CPSU_GREEN)
    c.setLineWidth(1.5)
    c.line(1*inch, height - 1.35*inch, width - 1*inch, height - 1.35*inch)
    
    # --- Body Content ---
    c.setFillColor(CPSU_GREEN)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width/2, height - 2*inch, "CLINICAL REFERRAL FORM")
    
    # Metadata
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 9)
    c.drawString(width - 3*inch, height - 2.4*inch, f"Date: {datetime.now().strftime('%B %d, %Y')}")
    c.drawString(width - 3*inch, height - 2.55*inch, f"Referral ID: {str(record_id)[:8].upper()}")
    
    # Patient Info Box
    c.setFillColor(CPSU_GREEN)
    c.rect(1*inch, height - 3.8*inch, width - 2*inch, 1*inch, fill=0, stroke=1)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(1.2*inch, height - 3*inch, "PATIENT INFORMATION")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1.2*inch, height - 3.25*inch, "Name:")
    c.drawString(1.2*inch, height - 3.45*inch, "Student ID:")
    c.drawString(1.2*inch, height - 3.65*inch, "College/Dept:")
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*inch, height - 3.25*inch, student_name)
    c.drawString(2.5*inch, height - 3.45*inch, student_sid or "N/A")
    c.drawString(2.5*inch, height - 3.65*inch, student_dept or "N/A")
    
    # Clinical Details
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(CPSU_GREEN)
    c.drawString(1*inch, height - 4.1*inch, "CLINICAL ASSESSMENT")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1.2*inch, height - 4.35*inch, "Primary Assessment:")
    c.drawString(1.2*inch, height - 4.55*inch, "Triage Level:")
    
    # Severity styling
    sev_text = ["Low", "Moderate", "High", "Emergency"][record.severity - 1]
    sev_color = [CPSU_GREEN, colors.orange, colors.red, colors.darkred][record.severity - 1]
    
    c.setFont("Helvetica", 10)
    c.drawString(2.5*inch, height - 4.35*inch, record.predicted_disease or "Undetermined")
    c.setFillColor(sev_color)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2.5*inch, height - 4.55*inch, sev_text)
    
    # Medical Notes
    c.setFillColor(CPSU_GREEN)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(1*inch, height - 5*inch, "PHYSICIAN'S NOTES & OBSERVATIONS")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    from reportlab.lib.utils import simpleSplit
    notes = record.staff_notes or "No additional clinical notes provided by the attending staff."
    lines = simpleSplit(notes, "Helvetica", 10, width - 2.2*inch)
    text_y = height - 5.25*inch
    for line in lines:
        c.drawString(1.1*inch, text_y, line)
        text_y -= 0.18*inch
        if text_y < 2.5*inch: break # Avoid overflow
        
    # --- Footer & Signatures ---
    c.setStrokeColor(colors.lightgrey)
    c.setLineWidth(0.5)
    c.line(1*inch, 2.8*inch, width - 1*inch, 2.8*inch)
    
    c.setFont("Helvetica", 9)
    c.drawString(1*inch, 2.4*inch, "Attending Health Officer:")
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1*inch, 1.8*inch, current_user.name.upper())
    c.setFont("Helvetica", 8)
    c.drawString(1*inch, 1.65*inch, "Registered Nurse / Medical Staff")
    c.line(1*inch, 1.78*inch, 3*inch, 1.78*inch)
    
    c.setFont("Helvetica", 9)
    c.drawString(width - 3*inch, 2.4*inch, "Authorized Stamp:")
    c.rect(width - 3*inch, 1.6*inch, 1.5*inch, 0.6*inch, stroke=1, fill=0) # Box for stamp
    
    # Bottom Footer
    c.setFillColor(colors.gray)
    c.setFont("Helvetica-Oblique", 8)
    c.drawCentredString(width/2, 0.5*inch, "This is a computer-generated official document from the CPSU HealthAI portal.")
    c.drawCentredString(width/2, 0.35*inch, "Verification of this referral can be requested at the University Health Services Office.")
    
    c.save()
    return FileResponse(pdf_path, filename=f"referral_{record_id}.pdf")

@router.get("/referral/{record_id}/xlsx")
async def export_referral_xlsx(
    record_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["staff", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
        
    result = await db.execute(select(SymptomRecord).where(SymptomRecord.id == record_id))
    record = result.scalars().first()
    
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
        
    data = {
        "Field": ["Referral ID", "Date", "Patient ID", "Predicted Disease", "Severity", "Staff Notes"],
        "Value": [
            str(record.id),
            str(record.created_at),
            str(record.student_id),
            record.predicted_disease,
            record.severity,
            record.staff_notes or ""
        ]
    }
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ReferralDetails')
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=referral_{record_id}.xlsx"}
    )
